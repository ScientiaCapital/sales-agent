"""
Excel Export Service

Professional Excel generation using openpyxl with multi-sheet support,
formatting, formulas, and charts.
"""

from typing import List, Dict, Any, Optional
from datetime import datetime
from io import BytesIO
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, PieChart, LineChart, Reference


logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Professional Excel exporter with openpyxl.

    Features:
    - Multi-sheet workbooks
    - Automatic column width adjustment
    - Header formatting with colors
    - Excel tables with filters
    - Number formatting for different data types
    - Optional charts (bar, pie, line)
    """

    def __init__(self, title: str):
        """
        Initialize Excel exporter.

        Args:
            title: Workbook title (used as filename base)
        """
        self.title = title
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def add_data_sheet(
        self,
        sheet_name: str,
        data: List[Dict[str, Any]],
        columns: List[str],
        add_table: bool = True,
        add_summary_row: bool = False,
        freeze_header: bool = True
    ) -> None:
        """
        Add a data sheet to the workbook.

        Args:
            sheet_name: Name of the sheet
            data: List of dictionaries containing row data
            columns: List of column names to include
            add_table: Whether to format as Excel table
            add_summary_row: Whether to add sum/count summary row
            freeze_header: Whether to freeze the header row
        """
        if not data:
            logger.warning(f"No data provided for sheet: {sheet_name}")
            return

        # Create sheet
        ws = self.workbook.create_sheet(title=sheet_name)

        # Write header row
        headers = [self._format_column_name(col) for col in columns]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data rows
        for row_num, row_data in enumerate(data, start=2):
            for col_num, col_name in enumerate(columns, start=1):
                value = row_data.get(col_name)

                # Format value based on type
                if isinstance(value, datetime):
                    formatted_value = value
                    ws.cell(row=row_num, column=col_num, value=formatted_value).number_format = 'yyyy-mm-dd hh:mm'
                elif isinstance(value, (int, float)):
                    ws.cell(row=row_num, column=col_num, value=value)
                    # Add number formatting for floats
                    if isinstance(value, float):
                        ws.cell(row=row_num, column=col_num).number_format = '#,##0.00'
                else:
                    ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')

        # Add summary row if requested
        if add_summary_row:
            self._add_summary_row(ws, len(data) + 2, columns, data)

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Add table formatting
        if add_table:
            self._format_as_table(ws, len(data), len(columns))

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = 'A2'

    def _format_column_name(self, column: str) -> str:
        """Format column name for display"""
        return column.replace('_', ' ').title()

    def _add_summary_row(
        self,
        ws,
        row_num: int,
        columns: List[str],
        data: List[Dict[str, Any]]
    ) -> None:
        """Add a summary row with totals/counts"""
        # First column: "Total" label
        cell = ws.cell(row=row_num, column=1, value="TOTAL")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        # Calculate sums for numeric columns
        for col_num, col_name in enumerate(columns, start=1):
            if col_num == 1:  # Skip first column (already set to "TOTAL")
                continue

            # Check if column has numeric data
            sample_value = data[0].get(col_name) if data else None

            if isinstance(sample_value, (int, float)):
                # Use SUM formula
                col_letter = get_column_letter(col_num)
                formula = f"=SUM({col_letter}2:{col_letter}{row_num - 1})"
                cell = ws.cell(row=row_num, column=col_num, value=formula)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

                if isinstance(sample_value, float):
                    cell.number_format = '#,##0.00'

    def _auto_adjust_columns(self, ws) -> None:
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_as_table(self, ws, num_rows: int, num_cols: int) -> None:
        """Format range as Excel table with filters"""
        # Define table range
        end_column = get_column_letter(num_cols)
        table_range = f"A1:{end_column}{num_rows + 1}"

        # Create table
        table = Table(displayName=ws.title.replace(' ', '_'), ref=table_range)

        # Apply table style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style

        ws.add_table(table)

    def add_chart_sheet(
        self,
        sheet_name: str,
        chart_type: str,
        data_sheet_name: str,
        x_column: str,
        y_columns: List[str],
        chart_title: str,
        x_axis_title: Optional[str] = None,
        y_axis_title: Optional[str] = None
    ) -> None:
        """
        Add a chart sheet to the workbook.

        Args:
            sheet_name: Name of the chart sheet
            chart_type: "bar", "line", or "pie"
            data_sheet_name: Name of the sheet containing the data
            x_column: Column name for x-axis (categories)
            y_columns: Column names for y-axis (values)
            chart_title: Chart title
            x_axis_title: X-axis label
            y_axis_title: Y-axis label
        """
        # Get data sheet
        data_ws = self.workbook[data_sheet_name]

        # Find column indices
        headers = [cell.value for cell in data_ws[1]]

        try:
            x_col_idx = headers.index(self._format_column_name(x_column)) + 1
            y_col_indices = [headers.index(self._format_column_name(col)) + 1 for col in y_columns]
        except ValueError as e:
            logger.error(f"Column not found: {e}")
            return

        # Create chart based on type
        if chart_type.lower() == "bar":
            chart = BarChart()
        elif chart_type.lower() == "line":
            chart = LineChart()
        elif chart_type.lower() == "pie":
            chart = PieChart()
        else:
            logger.error(f"Unsupported chart type: {chart_type}")
            return

        # Set chart properties
        chart.title = chart_title
        if hasattr(chart, 'x_axis') and x_axis_title:
            chart.x_axis.title = x_axis_title
        if hasattr(chart, 'y_axis') and y_axis_title:
            chart.y_axis.title = y_axis_title

        # Add data to chart
        max_row = data_ws.max_row

        # Categories (x-axis)
        cats = Reference(data_ws, min_col=x_col_idx, min_row=2, max_row=max_row)

        # Data series (y-axis)
        for y_col_idx in y_col_indices:
            data = Reference(data_ws, min_col=y_col_idx, min_row=1, max_row=max_row)
            chart.add_data(data, titles_from_data=True)

        if chart_type.lower() != "pie":
            chart.set_categories(cats)

        # Create new sheet for chart
        chart_ws = self.workbook.create_sheet(title=sheet_name)
        chart_ws.add_chart(chart, "A1")

    def add_summary_sheet(
        self,
        metadata: Dict[str, Any]
    ) -> None:
        """
        Add a summary/cover sheet with metadata.

        Args:
            metadata: Dictionary with report metadata
        """
        ws = self.workbook.create_sheet(title="Summary", index=0)

        # Title
        ws['A1'] = self.title
        ws['A1'].font = Font(size=20, bold=True, color="1F2937")
        ws.merge_cells('A1:D1')

        # Metadata
        row = 3
        for key, value in metadata.items():
            ws[f'A{row}'] = f"{key}:"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = str(value)
            row += 1

        # Auto-adjust columns
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40

    def export(self) -> BytesIO:
        """
        Export workbook to BytesIO buffer.

        Returns:
            BytesIO containing the Excel file
        """
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


# Convenience function for simple exports
def export_to_excel(
    data: List[Dict[str, Any]],
    columns: List[str],
    title: str,
    sheet_name: str = "Data"
) -> BytesIO:
    """
    Quick export function for simple Excel generation.

    Args:
        data: Data to export
        columns: Column names
        title: Workbook title
        sheet_name: Sheet name

    Returns:
        BytesIO containing Excel file
    """
    exporter = ExcelExporter(title=title)

    # Add metadata summary sheet
    exporter.add_summary_sheet({
        "Report Title": title,
        "Generated": datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'),
        "Total Records": len(data),
        "Columns": len(columns)
    })

    # Add data sheet
    exporter.add_data_sheet(
        sheet_name=sheet_name,
        data=data,
        columns=columns,
        add_table=True,
        add_summary_row=True,
        freeze_header=True
    )

    return exporter.export()
